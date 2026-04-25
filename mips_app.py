import streamlit as st
import pandas as pd

# ==============================
# Session State Defaults
# ==============================
st.session_state.setdefault("scored", False)
st.session_state.setdefault("results_df", None)
st.session_state.setdefault("quality_score", 0)
st.session_state.setdefault("quality_percent", 0)
st.session_state.setdefault("pi_percent", 0)
st.session_state.setdefault("final_score", 0)
st.session_state.setdefault("counted_measures", 0)
st.session_state.setdefault("max_points", 0)

st.set_page_config(page_title="MIPS Scoring Engine", layout="wide")
st.title("🏥 MIPS Scoring Engine")

# ==============================
# Helper Functions
# ==============================
def clean(val):
    """Normalize measure ID / collection type: uppercase string, strip whitespace."""
    if pd.isna(val):
        return ""
    return str(val).strip().upper()

def parse_range(val):
    """
    Parse decile range from benchmark:
    Accepts values like '10-20', '>=90', '<=50', or '--'. 
    Returns (low, high) numeric, or None for no data.
    """
    if pd.isna(val):
        return None
    val = str(val).strip()
    if val == "--" or val == "":
        return None
    # Remove percent sign if present
    val = val.replace("%", "")
    # Handle >= and <=
    if val.startswith(">="):
        try:
            low = float(val.replace(">=", "").strip())
        except:
            return None
        return (low, float("inf"))
    if val.startswith("<="):
        try:
            high = float(val.replace("<=", "").strip())
        except:
            return None
        return (float("-inf"), high)
    # Split hyphen range
    parts = val.split("-")
    if len(parts) == 2:
        try:
            low = float(parts[0].strip())
            high = float(parts[1].strip())
            return (low, high)
        except:
            return None
    # Single number
    try:
        v = float(val)
        return (v, v)
    except:
        return None

def get_score(rate, bench_row):
    """
    Compute achievement points for a measure.
    - rate: performance rate (percent)
    - bench_row: series containing Decile 1..10 ranges and 'Inverse' flag.
    Returns 0-10 points.
    """
    if rate is None or pd.isna(rate):
        return 0
    try:
        rate = float(rate)
    except:
        return 0
    inverse = str(bench_row.get('Inverse', '')).strip().upper() == 'YES'
    for i in range(1, 11):
        decile_val = bench_row.get(f'Decile {i}')
        rng = parse_range(decile_val)
        if rng:
            low, high = rng
            if inverse:
                # Inverse measure: better = lower rate
                if high <= rate <= low:
                    return i
            else:
                if low <= rate <= high:
                    return i
    return 0

def score_measures(perf_df, bench_df, small_practice=False):
    """
    Score each submitted measure against benchmarks.
    Returns DataFrame with columns: Measure ID, Title, Type, Rate, Score, Matched.
    """
    results = []
    for _, row in perf_df.iterrows():
        measure_id = clean(row.get('Measure ID'))
        coll_type = clean(row.get('Collection Type'))
        rate = row.get('Performance Rate')

        # Find matching benchmark row
        match = bench_df[
            (bench_df['Measure ID'] == measure_id) &
            (bench_df['Collection Type'] == coll_type)
        ]
        if match.empty:
            # No benchmark: 0 for large, 3 for small
            score = 3 if small_practice else 0
            results.append({
                'Measure ID': measure_id,
                'Measure Title': row.get('Measure Title', 'Unknown'),
                'Collection Type': coll_type,
                'Performance Rate': rate,
                'Score': score,
                'Matched': False
            })
            continue

        bench_row = match.iloc[0]
        score = get_score(rate, bench_row)
        # Apply 7-point cap if flagged
        if str(bench_row.get('Seven Point Cap', '')).strip().upper() == 'YES':
            score = min(score, 7)
        results.append({
            'Measure ID': measure_id,
            'Measure Title': bench_row.get('Measure Title', 'Unknown'),
            'Collection Type': coll_type,
            'Performance Rate': rate,
            'Score': score,
            'Matched': True
        })
    return pd.DataFrame(results)

# ==============================
# Sidebar Inputs
# ==============================
benchmark_file = st.sidebar.file_uploader("Upload Benchmark Excel", type=["xlsx", "xls"])
performance_file = st.sidebar.file_uploader("Upload Performance Excel", type=["xlsx", "xls"])
small_practice = st.sidebar.selectbox("Small Practice (15 or fewer clinicians)?", ["No", "Yes"])

# ==============================
# Main Scoring Logic
# ==============================
if benchmark_file and performance_file:
    # Read input files
    bench_df = pd.read_excel(benchmark_file)
    perf_df = pd.read_excel(performance_file)

    # Clean identifiers
    for col in ['Measure ID', 'Collection Type']:
        if col in bench_df.columns:
            bench_df[col] = bench_df[col].apply(clean)
        if col in perf_df.columns:
            perf_df[col] = perf_df[col].apply(clean)

    if st.button("🚀 Run Scoring"):
        sp_flag = (small_practice == "Yes")
        df_scores = score_measures(perf_df, bench_df, small_practice=sp_flag)

        # Sort by score and mark top 6 measures
        df_scores = df_scores.sort_values(by='Score', ascending=False).reset_index(drop=True)
        df_scores['Counted'] = False
        top_n = min(6, len(df_scores))
        df_scores.loc[:top_n-1, 'Counted'] = True
        
        # ✅ NEW: Dynamic denominator
        counted_measures = df_scores['Counted'].sum()
        max_points = counted_measures * 10

        # ==============================
        # CORRECT CMS QUALITY LOGIC
        # ==============================
        counted_measures = df_scores['Counted'].sum()
        max_points = counted_measures * 10  # dynamic max

        raw_quality = df_scores[df_scores['Counted']]['Score'].sum()

        # Cap at dynamic max (NOT always 60)
        quality_score = min(raw_quality, max_points)
        
        # Apply Small Practice Bonus ONLY if eligible
        if small_practice == "Yes":
            bonus = 6 if raw_quality > 0 else 0
        else:
            bonus = 0

        # Numerator can exceed 60
        numerator = raw_quality + bonus

        # Convert to %
        counted_measures = df_scores['Counted'].sum()

        # Avoid division by zero
        if counted_measures == 0:
            quality_percent = 0
        else:
            max_points = counted_measures * 10
            quality_percent = (numerator / max_points) * 100 if max_points > 0 else 0

        # 🔴 CRITICAL: CAP AT 100%
        quality_percent = min(quality_percent, 100)

        # Convert to weighted points (30%)
        quality_points = (quality_percent / 100) * 30

        # Save results in session state
        st.session_state.scored = True
        st.session_state.results_df = df_scores
        st.session_state.quality_score = raw_quality
        st.session_state.final_score = quality_points
        st.session_state.quality_percent = quality_percent
        st.session_state.numerator = numerator
        st.session_state.counted_measures = counted_measures
        st.session_state.max_points = max_points

# ==============================
# Display Results
# ==============================
if st.session_state.scored:
    
    df_res = st.session_state.results_df
    q_score = st.session_state.quality_score
    fin_score = st.session_state.final_score

    st.success("✅ Scoring Completed")

    st.dataframe(df_res, use_container_width=True)
    
    st.markdown(
    f"### Score (Top {st.session_state.counted_measures}): "
    f"{q_score} / {st.session_state.max_points}"
    )

    if small_practice == "Yes":
        bonus = st.session_state.numerator - q_score
        st.markdown(f"### Small Practice Bonus: +{bonus}")
        st.markdown(f"### Total (with bonus): {st.session_state.numerator} / {st.session_state.max_points}")
    else:
        st.markdown("### Small Practice Bonus: Not Applicable (Large Practice)")

    # ✅ OUTSIDE else
    st.markdown(f"### Quality Performance: {round(st.session_state.quality_percent,2)}% (Capped at 100%)")
    
    counted_measures = st.session_state.results_df['Counted'].sum()
    max_points = counted_measures * 10
    
    st.markdown(f"**Measures Counted:** {st.session_state.counted_measures}")
    st.markdown(f"**Max Possible Points:** {st.session_state.max_points}")
    
    st.markdown(f"### ✅ Quality Points (30% Weight): {round(fin_score,2)} / 30")

    # Display reweight scenarios
    if small_practice == "No":
        st.markdown("### 🏢 Large Practice Reweighting Scenarios")
    else:
        st.markdown("### ⚖️ Small Practice Reweighting Scenarios")
    
    base = st.session_state.quality_percent / 100  # fraction of quality potential
    # Define weight scenarios [Q%, C%, IA%, PI%]
   
    if small_practice == "No":  # 👉 Large Practice

        scenarios = {
            # Baseline
            "No Reweighting Needed": [30, 30, 15, 25],

            # Reweight 1
            "No Cost": [55, 0, 15, 30],
            "No PI": [55, 30, 15, 0],
            "No Quality": [0, 30, 15, 55],
            "No IA": [45, 30, 0, 25],

            # Reweight 2
            "No Cost & No PI": [85, 0, 15, 0],
            "No Cost & No Quality": [0, 0, 15, 85],
            "No Cost & No IA": [70, 0, 0, 30],
            "No PI & No Quality": [0, 50, 50, 0],
            "No PI & No IA": [70, 30, 0, 0],
            "No Quality & No IA": [0, 30, 0, 70],
        }
    
    else:  # 👉 Small Practice
        scenarios = {
            "No Reweight (baseline)": [30, 30, 15, 25],
            "No Cost":               [55,  0, 15, 30],
            "No PI (*Small Practice)":[40, 30, 30,  0],
            "No Quality":            [ 0, 30, 15, 55],
            "No IA":                 [45, 30,  0, 25],
            "No Cost & No PI":       [50,  0, 50,  0],
            "No Cost & No Quality":  [ 0,  0, 15, 85],
            "No Cost & No IA":       [70,  0,  0, 30],
            "No PI & No Quality":    [ 0, 50, 50,  0],
            "No PI & No IA":         [70, 30,  0,  0],
            "No Quality & No IA":    [ 0, 30,  0, 70]
        }
    
    reweight_df = pd.DataFrame.from_dict(
        scenarios, orient='index', columns=['Quality%', 'Cost%', 'IA%', 'PI%']
    )
    reweight_df['Quality Score'] = round(base * reweight_df['Quality%'], 2)
    st.table(reweight_df)

    # ==============================
    # PI Score Calculator (simplified)
    # ==============================
    st.markdown("## 💻 PI Score Calculator")

    # ROW 1 — eRx
    col1, col2 = st.columns(2)
    with col1:
        ep_attest = st.selectbox("e-Prescribing Reported?", ["Yes", "No"], key="ep_attest")
    with col2:
        ep_rate = st.number_input("eRx Performance (%)", 0, 100, key="ep_rate")

    # ROW 2 — PDMP (FIXED)
    col1, col2 = st.columns(2)
    with col1:
        pdmp_attest = st.selectbox("Query PDMP Reported?", ["Yes", "No"], key="pdmp_attest")
    with col2:
        st.markdown("<span style='color:gray'>PDMP is Yes/No only</span>", unsafe_allow_html=True)

    # ROW 3 — HIE Send
    col1, col2 = st.columns(2)
    with col1:
        hie_send_attest = st.selectbox("HIE – Send Reported?", ["Yes", "No"], key="hie_send_attest")
    with col2:
        hie_send_rate = st.number_input("HIE Send (%)", 0, 100, key="hie_send_rate")

    # ROW 4 — HIE Receive
    col1, col2 = st.columns(2)
    with col1:
        hie_recv_attest = st.selectbox("HIE – Receive Reported?", ["Yes", "No"], key="hie_recv_attest")
    with col2:
        hie_recv_rate = st.number_input("HIE Receive (%)", 0, 100, key="hie_recv_rate")

    # ROW — PEA1 (ADD THIS)
    col1, col2 = st.columns(2)
    with col1:
        pea_attest = st.selectbox("PEA1 Reported?", ["Yes", "No"], key="pea_attest")
    with col2:
        pea_rate = st.number_input("PEA1 Performance (%)", 0.0, 100.0, key="pea_rate")

    # ROW 5B — Public Health BONUS (ADD THIS HERE)
    col1, col2 = st.columns(2)
    with col1:
        ph_bonus = st.selectbox(
            "Public Health Bonus (Registry / Syndromic Surveillance Active Engagement)?",
            ["No", "Yes"],
            key="ph_bonus"
        )
    with col2:
        st.markdown("<span style='color:gray'>+5 bonus points (Active Engagement required)</span>", unsafe_allow_html=True)

    # ROW 6 — IR/eCR (FIXED)
    col1, col2 = st.columns(2)
    with col1:
        ir_attest = st.selectbox("IR/eCR Reported?", ["Yes", "No"], key="ir_attest")
    with col2:
        st.markdown("<span style='color:gray'>IR/eCR is Yes/No only</span>", unsafe_allow_html=True)

    # ==============================
    # CMS-ACCURATE PI CALCULATION
    # ==============================

    weights = {
        "ep": 10,
        "pdmp": 10,
        "hie_send": 15,
        "hie_recv": 15,
        "ph": 40,
        "ir": 25
    }

    earned = 0
    available = 0

    # e-Prescribing
    ep_score = 0
    if ep_attest == "Yes":
        ep_score = (ep_rate / 100) * weights["ep"]
        earned += ep_score
        available += weights["ep"]

    # PDMP
    pdmp_score = 0
    if pdmp_attest == "Yes":
        pdmp_score = weights["pdmp"]
        earned += pdmp_score
        available += weights["pdmp"]

    # HIE Send
    hie_send_score = 0
    if hie_send_attest == "Yes":
        hie_send_score = (hie_send_rate / 100) * weights["hie_send"]
        earned += hie_send_score
        available += weights["hie_send"]

    # HIE Receive
    hie_recv_score = 0
    if hie_recv_attest == "Yes":
        hie_recv_score = (hie_recv_rate / 100) * weights["hie_recv"]
        earned += hie_recv_score
        available += weights["hie_recv"]

    # ==============================
    # PEA1 (DYNAMIC CMS LOGIC)
    # ==============================

    pea_score = 0

    if pea_attest == "No":
        pea_score = 0
    else:
        rate = pea_rate / 100

        if (ep_attest=="No" and hie_send_attest=="No" and hie_recv_attest=="No" and ir_attest=="No"):
            pea_score = rate * 100

        elif (ep_attest=="Yes" and hie_send_attest=="No" and hie_recv_attest=="No" and ir_attest=="Yes"):
            pea_score = rate * 80

        elif (ep_attest=="Yes" and hie_send_attest=="No" and hie_recv_attest=="No" and ir_attest=="No"):
            pea_score = rate * 90

        elif (ep_attest=="Yes" and hie_send_attest=="No" and hie_recv_attest=="Yes" and ir_attest=="Yes"):
            pea_score = rate * 60

        elif (ep_attest=="Yes" and hie_send_attest=="Yes" and hie_recv_attest=="Yes" and ir_attest=="No"):
            pea_score = rate * 50

        elif (ep_attest=="No" and hie_send_attest=="No" and hie_recv_attest=="Yes" and ir_attest=="No"):
            pea_score = rate * 75

        elif (ep_attest=="Yes" and hie_send_attest=="No" and hie_recv_attest=="Yes" and ir_attest=="No"):
            pea_score = rate * 70

        elif (ep_attest=="No" and hie_send_attest=="No" and hie_recv_attest=="No" and ir_attest=="Yes"):
            pea_score = rate * 90

        elif (ep_attest=="No" and hie_send_attest=="No" and hie_recv_attest=="Yes" and ir_attest=="Yes"):
            pea_score = rate * 65

        else:
            pea_score = rate * 25

    # IR/eCR
    ir_score = 0
    if ir_attest == "Yes":
        ir_score = weights["ir"]
        earned += ir_score
        available += weights["ir"]
        
       

    # ==============================
    # BONUS (CRITICAL FIX)
    # ==============================
    bonus_points = 0

    if ph_bonus == "Yes":
        bonus_points += 5

    # ==============================
    # NORMALIZATION (CRITICAL FIX)
    # ==============================
    if available == 0:
        pi_percent = 0
    else:
        pi_percent = (earned / available) * 100

    # Apply bonus AFTER normalization
    pi_percent = pi_percent + bonus_points

    # 🔴 CAP AT 100%
    pi_percent = min(pi_percent, 100)
    st.session_state.pi_percent = pi_percent

    # Convert to MIPS points (25%)
    pi_weight_score = round((pi_percent / 100) * 25, 2)

    # ==============================
    # TABLE DISPLAY
    # ==============================
    pi_df = pd.DataFrame({
        "Measure": [
            "e-Prescribing", "PDMP Query", "HIE Send",
            "HIE Receive", "PEA1", "IR/eCR", "PH Bonus"
        ],
        "Attested": [
            ep_attest, pdmp_attest, hie_send_attest,
            hie_recv_attest, pea_attest, ir_attest, ph_bonus
        ],
        "Perf Rate (%)": [
            ep_rate, "-", hie_send_rate,
            hie_recv_rate, pea_rate, "-", "-"
        ],
        "Score": [
            round(ep_score,2),
            pdmp_score,
            round(hie_send_score,2),
            round(hie_recv_score,2),
            round(pea_score,2),
            ir_score,
            bonus_points
        ]
    })

    st.dataframe(pi_df, use_container_width=True)

    # ==============================
    # OUTPUT
    # ==============================
    col3, col4 = st.columns(2)

    with col3:
        st.markdown(f"**Earned Points:** {round(earned,2)} / {available}")

    with col4:
        st.markdown(f"**PI Score (CAPPED): {round(pi_percent,2)}% → {pi_weight_score} / 25**")



# ==============================
# 📊 MIPS FINAL DASHBOARD (AFTER PI)
# ==============================

st.markdown("## 📊 MIPS Final Score Scenarios Dashboard")

# ==============================
# Manual Inputs (IA, Cost, Bonus)
# ==============================
col1, col2, col3 = st.columns(3)

with col1:
    ia_score = st.number_input("IA Completion (%)", 0.0, 100.0, value=0.0, key="ia_score_final")

with col2:
    cost_score = st.number_input("Cost Performance (%)", 0.0, 100.0, value=0.0, key="cost_score_final")

with col3:
    complex_bonus = st.number_input("Complex Patient Bonus", 0.0, 10.0, value=0.0)

# ==============================
# Base Scores
# ==============================
base_quality = st.session_state.get("quality_percent", 0) / 100
base_pi = st.session_state.get("pi_percent", 0) / 100

# ==============================
# Scenario Calculator
# ==============================
def calc_total(q_w, c_w, ia_w, pi_w):
    quality_part = base_quality * q_w
    pi_part = base_pi * pi_w
    ia_part = (ia_score / 100) * ia_w
    cost_part = (cost_score / 100) * c_w if c_w > 0 else 0

    total = quality_part + pi_part + ia_part + cost_part + complex_bonus

    return (
        round(quality_part, 3),
        q_w,

        round(pi_part, 3),
        pi_w,

        round(ia_part, 3),
        ia_w,

        round(cost_part, 3),
        c_w,

        round(total, 3)
    )

# ==============================
# Scenarios (MATCH YOUR TABLE)
# ==============================
if small_practice == "No":  # 🔵 LARGE PRACTICE

    scenarios = {
        # Baseline
        "No Reweighting Needed": [30, 30, 15, 25],

        # Reweight 1
        "No Cost": [55, 0, 15, 30],
        "No PI": [55, 30, 15, 0],
        "No Quality": [0, 30, 15, 55],
        "No IA": [45, 30, 0, 25],

        # Reweight 2
        "No Cost and no PI": [85, 0, 15, 0],
        "No Cost and no Quality": [0, 0, 15, 85],
        "No Cost and no IA": [70, 0, 0, 30],
        "No PI and no Quality": [0, 50, 50, 0],
        "No PI and no IA": [70, 30, 0, 0],
        "No Quality and no IA": [0, 30, 0, 70],
    }

else:  # 🟢 SMALL PRACTICE

    scenarios = {
        # Baseline
        "No Reweighting Needed": [30, 30, 15, 25],

        # Reweight 1
        "No Cost": [55, 0, 15, 30],
        "No PI*": [40, 30, 30, 0],
        "No Quality": [0, 30, 15, 55],
        "No IA": [45, 30, 0, 25],

        # Reweight 2
        "No Cost and no PI*": [50, 0, 50, 0],
        "No Cost and no Quality": [0, 0, 15, 85],
        "No Cost and no IA": [70, 0, 0, 30],
        "No PI and no Quality": [0, 50, 50, 0],
        "No PI and no IA": [70, 30, 0, 0],
        "No Quality and no IA": [0, 30, 0, 70],
    }

# ==============================
# Grouping Helper (ADD HERE)
# ==============================
def get_group(name):
    if "and" in name:
        return "Reweight 2"
    elif name == "No Reweighting Needed":
        return "Baseline"
    else:
        return "Reweight 1"


# ==============================
# Build Dashboard Table
# ==============================
rows = []

for name, (q_w, c_w, ia_w, pi_w) in scenarios.items():
    q, q_w, pi, pi_w, ia, ia_w, cost, c_w, total = calc_total(q_w, c_w, ia_w, pi_w)

    rows.append({
        "Group": get_group(name),
        "Scenario": name,
        "Quality": f"{q} / {q_w}" if q_w > 0 else "—",
        "PI": f"{round(pi,2)} / {pi_w}" if pi_w > 0 else "—",
        "IA": f"{round(ia,2)} / {ia_w}" if ia_w > 0 else "—",
        "Cost": (
            f"{round(cost,2)} / {c_w}" if c_w > 0 
            else ("To be calculated by CMS" if name=="No Reweighting Needed" else "—")
        ),
        "Complex Patients": complex_bonus,
        "Total": total
    })

dashboard_df = pd.DataFrame(rows)

# ==============================
# Display Table
# ==============================
def highlight_total(val):
    try:
        if float(val) >= 75:
            return "background-color: #90EE90; font-weight: bold;"  # light green
    except:
        pass
    return ""

dashboard_df = pd.DataFrame(rows)

# Sort by group for clean layout
dashboard_df = dashboard_df.sort_values(by=["Group", "Scenario"])

styled_df = (
    dashboard_df.style
    .map(your_function)
    .apply(highlight_total, subset=["Total"])
)

st.dataframe(styled_df, use_container_width=True)



# ==============================
# Download Results
# ==============================
from io import BytesIO
from openpyxl.styles import PatternFill

if st.session_state.scored:

    st.markdown("### 📥 Download Results")

    output = BytesIO()

    # ✅ ADD THIS HERE (STEP 1)
    quality_summary = pd.DataFrame({
        "Metric": [
            "Quality Score",
            "Small Practice Bonus",
            "Total (with bonus)",
            "Max Possible Points",
            "Quality %",
            "Quality Points (30%)"
        ],
        "Value": [
            st.session_state.quality_score,
            st.session_state.numerator - st.session_state.quality_score,
            st.session_state.numerator,
            st.session_state.max_points,
            f"{round(st.session_state.quality_percent,2)}%",
            round(st.session_state.final_score,2)
        ]
    })

    pi_summary = pd.DataFrame({
        "Metric": [
            "Earned Points",
            "Available Points",
            "PI Score (%)",
            "PI Points (25%)"
        ],
        "Value": [
            round(earned, 2),
            available,
            f"{round(pi_percent, 2)}%",
            pi_weight_score
        ]
    })

    # Ensure numeric
    dashboard_df["Total"] = pd.to_numeric(dashboard_df["Total"], errors="coerce")

    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        # =========================
        # QUALITY DETAILS (FIRST)
        # =========================
        df_res.to_excel(
            writer,
            index=False,
            sheet_name='MIPS Report',
            startrow=0
        )

        # =========================
        # QUALITY SUMMARY (AFTER DETAILS)
        # =========================
        summary_start = len(df_res) + 3

        quality_summary.to_excel(
            writer,
            index=False,
            sheet_name='MIPS Report',
            startrow=summary_start
        )

        # =========================
        # PI DETAILS (FIRST)
        # =========================
        pi_start = summary_start + len(quality_summary) + 3

        pi_df.to_excel(
            writer,
            index=False,
            sheet_name='MIPS Report',
            startrow=pi_start
        )

        # =========================
        # PI SUMMARY (AFTER DETAILS)
        # =========================
        pi_summary_start = pi_start + len(pi_df) + 3

        pi_summary.to_excel(
            writer,
            index=False,
            sheet_name='MIPS Report',
            startrow=pi_summary_start
        )

        # =========================
        # FINAL SCENARIOS (SHIFT DOWN)
        # =========================
        dashboard_start = pi_summary_start + len(pi_summary) + 4

        dashboard_df.to_excel(
            writer,
            index=False,
            sheet_name='MIPS Report',
            startrow=dashboard_start
        )

        # =========================
        # LABELS
        # =========================
        ws = writer.sheets['MIPS Report']

        ws.cell(row=1, column=1).value = "QUALITY MEASURE DETAILS"
        ws.cell(row=summary_start+1, column=1).value = "QUALITY SUMMARY"
        ws.cell(row=pi_start+1, column=1).value = "PI DETAILS"
        ws.cell(row=pi_summary_start+1, column=1).value = "PI SUMMARY"
        ws.cell(row=dashboard_start+1, column=1).value = "FINAL SCENARIOS"

        # Freeze (adjusted for summary at top)
        ws.freeze_panes = "A2"

        # Auto width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Highlight ≥75
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        for r in range(1, ws.max_row + 1):
            headers = [cell.value for cell in ws[r]]

            if "Total" in headers:
                total_col_idx = headers.index("Total") + 1

                for row in range(r + 1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=total_col_idx)
                    try:
                        if float(cell.value) >= 75:
                            cell.fill = green_fill
                    except:
                        pass
                break

    # ✅ IMPORTANT: move this OUTSIDE the writer block
    output.seek(0)

    st.download_button(
        label="Download Full MIPS Report",
        data=output.getvalue(),
        file_name="MIPS_Scoring_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
